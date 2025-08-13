#! /usr/bin/env nix-shell
#! nix-shell -i python3 -p python3Packages.openpyxl

import argparse
from dataclasses import dataclass, asdict
import json
import logging
from openpyxl import load_workbook


logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


@dataclass
class Station:
    운영기관명: str
    노선명: str
    역번호: str
    한글역명: str
    영어역명: str
    로마자역명: str | None
    일본어역명: str | None
    중국어간체역명: str | None
    중국어번체역명: str | None
    부역명: str | None
    환승: bool
    환승노선명: str | None
    경도: float
    위도: float
    도로명주소: str
    전화번호: str | None
    신설일자: str | None
    데이터기준일자: str | None


def parse_cell(value, to):
    # Convert cell value to the specified type (or None)

    if isinstance(value, str):
        value = value.strip()
        if value == "" or value == "-" or value == "해당없음":
            value = None

    if value is None:
        if to is bool:
            return False
        return None
    if to is bool:
        if isinstance(value, str):
            return value in ("가능", "O", "있음")
        logger.warning(f"Cannot convert {value} to bool, returning False.")
        return False
    if to is str:
        return str(value)
    if to is float:
        return float(value)
    if to is int:
        return int(value)

    raise ValueError(f"Unsupported type: {to}")


def main() -> None:
    parser = argparse.ArgumentParser(
        prog="xlsx2geojson",
        description="Convert 전국 도시광역철도 역사정보 xlsx to GeoJSON.",
    )
    parser.add_argument("filename", type=str)
    args = parser.parse_args()
    filename = args.filename

    logger.info(f"File: {filename}")

    wb = load_workbook(filename, read_only=True, data_only=True)
    ws = wb.active

    if ws is None:
        logger.error(f"Error: No active worksheet found in {filename}")
        exit(1)

    features: list[Station] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        # Parse each row into a Station object.
        station = Station(
            parse_cell(row[0], str),  # type: ignore 운영기관명
            parse_cell(row[1], str),  # type: ignore 노선명
            parse_cell(row[3], str),  # type: ignore 역번호
            parse_cell(row[4], str),  # type: ignore 한글역명
            parse_cell(row[5], str),  # type: ignore 영어역명
            parse_cell(row[6], str),  # type: ignore 로마자역명
            parse_cell(row[7], str),  # type: ignore 일본어역명
            parse_cell(row[8], str),  # type: ignore 중국어간체역명
            parse_cell(row[9], str),  # type: ignore 중국어번체역명
            parse_cell(row[10], str),  # type: ignore 부역명
            parse_cell(row[11], bool),  # type: ignore 환승
            parse_cell(row[12], str),  # type: ignore 환승노선명
            parse_cell(row[18], float),  # type: ignore 경도
            parse_cell(row[19], float),  # type: ignore 위도
            parse_cell(row[21], str),  # type: ignore 도로명주소
            parse_cell(row[22], str),  # type: ignore 전화번호
            parse_cell(row[23], str),  # type: ignore 신설일자
            parse_cell(row[27], str),  # type: ignore 데이터기준일자
        )
        # Ensure that the station has a valid name and coordinates
        if station.경도 is None or station.위도 is None:
            logger.warning(
                f"Skipping station with missing coordinates: {station.한글역명}"
            )
            continue
        features.append(station)

    logger.info(f"Parsed {len(features)} stations.")

    geojson = {
        "type": "FeatureCollection",
        "features": [
            {
                "type": "Feature",
                "geometry": {
                    "type": "Point",
                    "coordinates": [feature.경도, feature.위도],
                },
                "properties": asdict(feature),
            }
            for feature in features
        ],
    }

    with open(
        "railway-stations-south-korea.geojson", mode="w", encoding="utf-8"
    ) as file:
        file.write(json.dumps(geojson, ensure_ascii=False, allow_nan=False))


if __name__ == "__main__":
    main()
